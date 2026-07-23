"""
Day 20 deliverable: unit tests for the peer comparison Excel export.
Run with: python -m pytest tests/kpi/test_peer_comparison_export.py -v
"""
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "src", "analytics"))

from export_peer_comparison import percentile_fill, GREEN_FILL, YELLOW_FILL, RED_FILL


def test_01_high_percentile_gets_green():
    assert percentile_fill(0.9) == GREEN_FILL


def test_02_exactly_75th_percentile_gets_green():
    assert percentile_fill(0.75) == GREEN_FILL


def test_03_mid_percentile_gets_yellow():
    assert percentile_fill(0.5) == YELLOW_FILL


def test_04_exactly_25th_percentile_gets_yellow():
    assert percentile_fill(0.25) == YELLOW_FILL


def test_05_low_percentile_gets_red():
    assert percentile_fill(0.1) == RED_FILL


def test_06_none_percentile_gets_no_fill():
    assert percentile_fill(None) is None


def test_07_output_file_has_11_sheets():
    from openpyxl import load_workbook
    wb = load_workbook("output/peer_comparison.xlsx")
    assert len(wb.sheetnames) == 11


def test_08_it_services_sheet_has_median_row():
    from openpyxl import load_workbook
    wb = load_workbook("output/peer_comparison.xlsx")
    ws = wb["IT Services"]
    # 5 companies + header + median row = row 7
    assert ws.cell(row=7, column=1).value == "MEDIAN"