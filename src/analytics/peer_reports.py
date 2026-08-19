"""Sprint 3 Day 19-20 — radar charts (PNG, via matplotlib since plotly is not
installable in this offline sandbox) + peer_comparison.xlsx export."""
import sys
import pathlib
import sqlite3
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))
from screener.engine import load_latest_universe
from analytics.peer import run_peer_percentiles

RADAR_AXES = ["return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
              "debt_to_equity", "free_cash_flow_cr", "pat_cagr_5yr", "revenue_cagr_5yr",
              "composite_quality_score"]
RADAR_LABELS = ["ROE", "ROCE", "NPM", "D/E", "FCF score", "PAT CAGR 5yr", "Revenue CAGR 5yr", "Composite"]

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _normalise_for_radar(df, axes):
    """0-100 min-max scaling per axis across the full universe, D/E inverted (lower better)."""
    out = pd.DataFrame(index=df.index)
    for ax in axes:
        s = df[ax].astype(float)
        lo, hi = s.min(), s.max()
        if pd.isna(lo) or pd.isna(hi) or hi == lo:
            out[ax] = 50.0
            continue
        scaled = (s - lo) / (hi - lo) * 100
        if ax == "debt_to_equity":
            scaled = 100 - scaled
        out[ax] = scaled
    return out


def make_radar_chart(company_row, peer_avg_row, ticker, out_path):
    values = [company_row.get(a, 0) or 0 for a in RADAR_AXES]
    avg_values = [peer_avg_row.get(a, 0) or 0 for a in RADAR_AXES]
    n = len(RADAR_AXES)
    angles = np.linspace(0, 2 * np.pi, n, endpoint=False).tolist()
    values += values[:1]
    avg_values += avg_values[:1]
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(5, 5), subplot_kw=dict(polar=True))
    ax.plot(angles, values, color="#1F4E78", linewidth=2, label=ticker)
    ax.fill(angles, values, color="#1F4E78", alpha=0.25)
    ax.plot(angles, avg_values, color="#C0504D", linewidth=1.5, linestyle="--", label="Peer group avg")
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(RADAR_LABELS, fontsize=8)
    ax.set_yticklabels([])
    ax.set_title(ticker, fontsize=12, weight="bold")
    ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1), fontsize=7)
    fig.tight_layout()
    fig.savefig(out_path, dpi=110)
    plt.close(fig)


def run_radar_charts(universe, peer_groups):
    pathlib.Path("reports/radar_charts").mkdir(parents=True, exist_ok=True)
    radar_df = _normalise_for_radar(universe, RADAR_AXES)
    radar_df["company_id"] = universe["company_id"].values

    n_generated = 0
    for group_name, members in peer_groups.groupby("peer_group_name"):
        member_ids = members["company_id"].tolist()
        grp = radar_df[radar_df["company_id"].isin(member_ids)]
        if grp.empty:
            continue
        peer_avg = grp[RADAR_AXES].mean()
        for _, row in grp.iterrows():
            make_radar_chart(row, peer_avg, row["company_id"],
                              f"reports/radar_charts/{row['company_id']}_radar.png")
            n_generated += 1

    # companies with no peer group -> standalone chart vs Nifty100 average
    universe_avg = radar_df[RADAR_AXES].mean()
    no_group_ids = set(universe["company_id"]) - set(peer_groups["company_id"])
    for cid in no_group_ids:
        row = radar_df[radar_df["company_id"] == cid]
        if row.empty:
            continue
        make_radar_chart(row.iloc[0], universe_avg, cid,
                          f"reports/radar_charts/{cid}_radar.png")
        n_generated += 1
    print(f"Radar charts generated: {n_generated}")
    return n_generated


def export_peer_comparison_xlsx(universe, peer_groups, percentiles_df):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    display_metrics = [m for m, _ in __import__("analytics.peer", fromlist=["METRICS"]).METRICS]

    for group_name, members in peer_groups.groupby("peer_group_name"):
        member_ids = members["company_id"].tolist()
        benchmark_ids = set(members[members["is_benchmark"] == 1]["company_id"])
        grp = universe[universe["company_id"].isin(member_ids)].copy()
        if grp.empty:
            continue
        ws = wb.create_sheet(group_name[:31])

        headers = ["company_id", "company_name"] + display_metrics + [f"{m}_pctile" for m in display_metrics]
        ws.append(headers)
        for c in ws[1]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT

        pct_pivot = percentiles_df[percentiles_df.peer_group_name == group_name].pivot_table(
            index="company_id", columns="metric", values="percentile_rank", aggfunc="last")

        for _, row in grp.iterrows():
            cid = row["company_id"]
            vals = [cid, row.get("company_name")]
            vals += [row.get(m) for m in display_metrics]
            pct_row = pct_pivot.loc[cid] if cid in pct_pivot.index else {}
            vals += [pct_row.get(m) if hasattr(pct_row, "get") else None for m in display_metrics]
            ws.append(vals)
            r = ws.max_row
            if cid in benchmark_ids:
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=r, column=ci).fill = GOLD
            else:
                pct_start = 2 + len(display_metrics) + 1
                for j, m in enumerate(display_metrics):
                    v = pct_row.get(m) if hasattr(pct_row, "get") else None
                    cell = ws.cell(row=r, column=pct_start + j)
                    if v is None or pd.isna(v):
                        continue
                    if v >= 0.75:
                        cell.fill = GREEN
                    elif v >= 0.25:
                        cell.fill = YELLOW
                    else:
                        cell.fill = RED

        # summary row: peer group median
        median_vals = ["MEDIAN", ""] + [grp[m].median() if m in grp else None for m in display_metrics] + [None] * len(display_metrics)
        ws.append(median_vals)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True, italic=True)

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14

    pathlib.Path("output").mkdir(exist_ok=True)
    wb.save("output/peer_comparison.xlsx")
    print(f"Saved output/peer_comparison.xlsx with {len(wb.sheetnames)} sheets")


if __name__ == "__main__":
    conn = sqlite3.connect("data/nifty100.db")
    universe = load_latest_universe(conn)
    peer_groups = pd.read_sql("SELECT * FROM peer_groups", conn)
    conn.close()
    percentiles_df, no_group = run_peer_percentiles()
    run_radar_charts(universe, peer_groups)
    export_peer_comparison_xlsx(universe, peer_groups, percentiles_df)
