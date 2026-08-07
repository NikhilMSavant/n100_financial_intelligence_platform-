"""
export_peer_comparison.py — Sprint 3 / Day 20
Generates output/peer_comparison.xlsx with 11 sheets (one per peer group),
percentile-rank colour coding, benchmark row highlighted, median summary row.
"""
import os
import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DB_PATH = os.path.join(ROOT, "db", "nifty100.db")
OUT_PATH = os.path.join(ROOT, "output", "peer_comparison.xlsx")

METRICS = [
    "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
    "debt_to_equity", "free_cash_flow_cr", "pat_cagr_5yr", "revenue_cagr_5yr",
    "eps_cagr_5yr", "interest_coverage", "asset_turnover",
]

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
BOLD = Font(bold=True)


def run():
    conn = sqlite3.connect(DB_PATH)
    fr = pd.read_sql("SELECT * FROM financial_ratios", conn)
    idx = fr.groupby("company_id")["year"].idxmax()
    latest = fr.loc[idx].reset_index(drop=True)
    companies = pd.read_sql("SELECT company_id, company_name FROM companies", conn)
    peer_groups = pd.read_sql("SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", conn)
    pct = pd.read_sql("SELECT * FROM peer_percentiles", conn)

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for group_name, members in peer_groups.groupby("peer_group_name"):
            ids = members["company_id"].tolist()
            g = latest[latest.company_id.isin(ids)].merge(companies, on="company_id", how="left")
            g = g.merge(members[["company_id", "is_benchmark"]], on="company_id", how="left")

            cols = ["company_id", "company_name"] + METRICS
            sheet_df = g[cols].copy()

            # attach percentile columns
            grp_pct = pct[pct.peer_group_name == group_name]
            for m in METRICS:
                pm = grp_pct[grp_pct.metric == m].set_index("company_id")["percentile_rank"]
                sheet_df[f"{m}_pctile"] = sheet_df["company_id"].map(pm)

            median_row = {"company_id": "", "company_name": "PEER GROUP MEDIAN"}
            for m in METRICS:
                median_row[m] = sheet_df[m].median()
                median_row[f"{m}_pctile"] = None
            sheet_df = pd.concat([sheet_df, pd.DataFrame([median_row])], ignore_index=True)

            sheet_name = group_name[:31]
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            header = {cell.value: idx2 + 1 for idx2, cell in enumerate(ws[1])}

            n_rows = len(sheet_df)
            for m in METRICS:
                pcol = header.get(f"{m}_pctile")
                if not pcol:
                    continue
                for r in range(2, n_rows + 1):  # last row is median summary
                    if r == n_rows + 1:
                        continue
                    cell = ws.cell(row=r, column=pcol)
                    v = cell.value
                    if v is None or not isinstance(v, (int, float)):
                        continue
                    if v >= 0.75:
                        cell.fill = GREEN
                    elif v >= 0.25:
                        cell.fill = YELLOW
                    else:
                        cell.fill = RED

            # highlight benchmark row(s)
            bench_ids = set(g[g.is_benchmark == 1]["company_id"])
            for r, cid in enumerate(sheet_df["company_id"], start=2):
                if cid in bench_ids:
                    for c in range(1, len(sheet_df.columns) + 1):
                        ws.cell(row=r, column=c).fill = GOLD
            # bold the median summary row
            for c in range(1, len(sheet_df.columns) + 1):
                ws.cell(row=n_rows + 1, column=c).font = BOLD

            for i, col in enumerate(sheet_df.columns, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(col)) + 2)

    print(f"Wrote {OUT_PATH} with {peer_groups['peer_group_name'].nunique()} sheets")
    conn.close()


if __name__ == "__main__":
    run()
