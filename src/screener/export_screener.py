"""
export_screener.py — Sprint 3 / Day 17
Generates output/screener_output.xlsx: one sheet per preset, 20 KPI columns,
sorted by composite score descending, green/red fill vs the preset threshold.
"""
import os
import sqlite3
import sys
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from engine import load_config, build_universe, apply_filters
from composite_score import compute as compute_composite

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DB_PATH = os.path.join(ROOT, "db", "nifty100.db")
OUT_PATH = os.path.join(ROOT, "output", "screener_output.xlsx")

DISPLAY_COLS = [
    "company_id", "company_name", "broad_sector", "composite_quality_score",
    "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
    "operating_profit_margin_pct", "debt_to_equity", "interest_coverage", "icr_label",
    "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "pe_ratio", "pb_ratio", "dividend_yield_pct", "dividend_payout_ratio_pct",
    "asset_turnover", "sales",
]

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# metric-key -> (dataframe column, comparison) for cell colour-coding
FILTER_COL_MAP = {
    "roe_min": ("return_on_equity_pct", "min"), "de_max": ("debt_to_equity", "max"),
    "fcf_min": ("free_cash_flow_cr", "min"), "revenue_cagr_5yr_min": ("revenue_cagr_5yr", "min"),
    "pat_cagr_5yr_min": ("pat_cagr_5yr", "min"), "opm_min": ("operating_profit_margin_pct", "min"),
    "pe_max": ("pe_ratio", "max"), "pb_max": ("pb_ratio", "max"),
    "dividend_yield_min": ("dividend_yield_pct", "min"), "sales_min": ("sales", "min"),
}


def run():
    conn = sqlite3.connect(DB_PATH)
    config = load_config()
    universe = build_universe(conn)
    universe = compute_composite(universe, sector_relative=False)

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for key, preset in config["presets"].items():
            result = apply_filters(universe, preset["filters"], config)
            result = result.sort_values("composite_quality_score", ascending=False)
            cols = [c for c in DISPLAY_COLS if c in result.columns]
            sheet_df = result[cols]
            sheet_name = preset["label"][:31]
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            for metric_key, threshold in preset["filters"].items():
                if metric_key not in FILTER_COL_MAP:
                    continue
                col_name, cmp = FILTER_COL_MAP[metric_key]
                if col_name not in header:
                    continue
                col_idx = header[col_name]
                for row_i, val in enumerate(sheet_df[col_name], start=2):
                    if pd.isna(val):
                        continue
                    passed = (val >= threshold) if cmp == "min" else (val <= threshold)
                    ws.cell(row=row_i, column=col_idx).fill = GREEN if passed else RED
            for i, col in enumerate(cols, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, len(col) + 2)

    print(f"Wrote {OUT_PATH}")
    conn.close()


if __name__ == "__main__":
    run()
