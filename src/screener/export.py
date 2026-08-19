"""Sprint 3 Day 17 — export screener_output.xlsx: one sheet per preset,
20 KPI columns, colour-coded pass/fail, sorted by rank metric."""
import sys
import pathlib
import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas as pd

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))
from screener.engine import run_screener

DISPLAY_COLS = [
    "company_id", "company_name", "broad_sector", "composite_quality_score",
    "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
    "debt_to_equity", "interest_coverage", "icr_label",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "free_cash_flow_cr", "fcf_yield_pct", "cfo_quality_label",
    "pe_ratio", "pb_ratio", "dividend_yield_pct", "sales", "net_profit",
]

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def field_passes(row, field, cond):
    val = row.get(field)
    if val is None or pd.isna(val):
        return None
    if field == "debt_to_equity" and row.get("broad_sector") == "Financials":
        return True
    if field == "interest_coverage" and row.get("icr_label") == "Debt Free":
        return True
    ok = True
    if "min" in cond:
        ok &= val >= cond["min"]
    if "max" in cond:
        ok &= val <= cond["max"]
    if "equals" in cond:
        ok &= val == cond["equals"]
    return ok


def export():
    df, results, config = run_screener()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for key, preset in config["presets"].items():
        res = results[key]
        cols = [c for c in DISPLAY_COLS if c in res.columns]
        sheet_df = res[cols].copy()
        ws = wb.create_sheet(preset["label"][:31])

        ws.append(cols)
        for c in ws[1]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT

        filters = preset["filters"] if key != "turnaround_watch" else {}
        for _, row in sheet_df.iterrows():
            ws.append([row[c] if pd.notna(row[c]) else None for c in cols])
            r = ws.max_row
            for ci, colname in enumerate(cols, start=1):
                if colname in filters:
                    passed = field_passes(row, colname, filters[colname])
                    if passed is True:
                        ws.cell(row=r, column=ci).fill = GREEN
                    elif passed is False:
                        ws.cell(row=r, column=ci).fill = RED
        for i, col in enumerate(cols, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(12, len(col) + 2)

    pathlib.Path("output").mkdir(exist_ok=True)
    wb.save("output/screener_output.xlsx")
    print("Saved output/screener_output.xlsx with", len(wb.sheetnames), "sheets")


if __name__ == "__main__":
    export()
